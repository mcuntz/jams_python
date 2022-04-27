#!/usr/bin/env python
from __future__ import division, absolute_import, print_function
from warnings import warn, filterwarnings
filterwarnings("default", category=DeprecationWarning)
import numpy as np


__all__ = ['xread', 'xlsread', 'xlsxread']


def xread(infile, sheet=None, nc=0, cname=None, snc=0, sname=None,
          skip=0, cskip=0, hskip=0, hstrip=True,
          squeeze=False, reform=False, transpose=False,
          fill=False, fill_value=0, sfill_value='', header=False):
    """
    Read from Excel file numbers into 2D float array as well as characters into
    2D string array.

    This routine is analog to fsread but for Excel files.


    Definition
    ----------
    def xread(infile, sheet=None, nc=0, cname=None, snc=0, sname=None,
              skip=0, cskip=0, hskip=0, hstrip=True,
              squeeze=False, reform=False, transpose=False,
              fill=False, fill_value=0, sfill_value='', header=False):

    Input
    -----
    infile         source file name

    Optional Input Parameters
    -------------------------
    sheet        Name or number of Excel sheet (default: first sheet)
    nc           number of columns to be read into float array
                 (default: all (nc<=0))
                 nc can be a vector of column indeces,
                 starting with 0; cskip will be ignored then.
                 if snc!=0: nc must be iterable or -1 for all other columns.
    cname        float columns can alternatively be chosen by the values in the
                 first header line; must be iterable with strings.
    snc          number of columns to be read into string array
                 (default: none (snc=0))
                 snc can be a vector of column indeces,
                 starting with 0; cskip will be ignored then.
                 if nc!=0: snc must be iterable or -1 for all other columns.
    sname        string columns can alternatively be chosen by the values in
                 the first header line; must be iterable with strings.
    skip         number of lines to skip at the beginning of file (default: 0)
    cskip        number of columns to skip at the beginning of each line
                 (default: 0)
    hskip        number of lines in skip that do not belong to header
                 (default: 0)
    hstrip       If true strip header cells to match with cname and sname
                 (default: True)
    fill_value   value to fill float array in empty cells or if not enough
                 columns per line and fill=True (default 0)
    sfill_value  value to fill in header and in string array in empty cells
                 or if not enough columns per line and fill=True (default '')

    Options
    -------
    squeeze      True:  2-dim array will be cleaned of degenerated
                        dimension, i.e. results in vector
                 False: array will be two-dimensional as read (default)
    reform       Same as squeeze.
    fill         True:  fills in fill_value if not enough columns in line
                 False: stops execution and returns None if not enough
                        columns in line (default)
    header       True:  header strings will be returned
                 False  numbers in file will be returned (default)
    transpose    True:  column-major format output(0:ncolumns,0:nlines)
                 False: row-major format output(0:nlines,0:ncolumns) (default)

    Output
    ------
    Depending on options:
        1 output:  2D-array of floats (snc=0)
        1 output:  list/2D-array of strings (nc=0 and snc!=0)
        2 outputs: 2D-array of floats, 2D-array of strings (nc!=0 and snc!=0)
        1 output:  list/string array of header
                   ((nc=0 or snc=0) and header=True)
        2 outputs: list/string array of header for float array,
                   list/string array of header for strarr
                   ((nc!=0 and snc!=0) and header=True)

    Restrictions
    ------------
    If header=True then skip is counterintuitive because it is
      actally the number of header rows to be read. This is to
      be able to have the exact same call of the function, once
      with header=False and once with header=True.

    Examples
    --------
    >>> from autostring import astr

    # xlrd
    >>> filename = 'test_readexcel.xls'
    >>> print(astr(xread(filename, skip=1), 1, pp=True))
    [['1.1' '1.2' '1.3' '1.4']
     ['2.1' '2.2' '2.3' '2.4']
     ['3.1' '3.2' '3.3' '3.4']
     ['4.1' '4.2' '4.3' '4.4']]
    >>> print(astr(xread(filename, skip=1, nc=[2], squeeze=True), 1, pp=True))
    ['1.3' '2.3' '3.3' '4.3']
    >>> print(astr(xread(filename, skip=1, cname=['head1', 'head2']),
    ...            1, pp=True))
    [['1.1' '1.2']
     ['2.1' '2.2']
     ['3.1' '3.2']
     ['4.1' '4.2']]
    >>> a, sa = xread(filename, sheet='Sheet3', nc=[1], snc=[0, 2], skip=1,
    ...               squeeze=True)
    >>> print(astr(a, 1, pp=True))
    ['1.2' '2.2' '3.2' '4.2']
    >>> print(sa)
    [['name1' 'name5']
     ['name2' 'name6']
     ['name3' 'name7']
     ['name4' 'name8']]
    >>> a, sa = xread(filename, sheet=2, cname='head2', snc=[0, 2], skip=1,
    ...               squeeze=True)
    >>> print(astr(a, 1, pp=True))
    ['1.2' '2.2' '3.2' '4.2']
    >>> print(sa)
    [['name1' 'name5']
     ['name2' 'name6']
     ['name3' 'name7']
     ['name4' 'name8']]
    >>> a, sa = xread(filename, sheet='Sheet2', cname=['head2', 'head4'],
    ...               snc=[0, 2], skip=1, fill=True, fill_value=-9,
    ...               sfill_value='-8')
    >>> print(astr(a, 1, pp=True))
    [['-9.0' ' 1.4']
     [' 2.2' ' 2.4']
     [' 3.2' ' 3.4']
     [' 4.2' ' 4.4']]
    >>> print(sa)
    [['1.1' '1.3']
     ['2.1' '2.3']
     ['3.1' '-8']
     ['4.1' '4.3']]
    >>> a, sa = xread(filename, sheet='Sheet2', cname=['head2', 'head4'],
    ...               snc=[0, 2], skip=1, header=True)
    >>> print(a)
    ['head2' 'head4']
    >>> print(sa)
    ['head1' 'head3']
    >>> print(xread(filename, sheet='Sheet2', skip=1, header=True))
    ['head1' 'head2' 'head3' 'head4']
    >>> a, sa = xread(filename, sheet='Sheet2', cname=[' head2', 'head4'],
    ...               snc=[0, 2], skip=1, fill=True, fill_value=-9,
    ...               sfill_value='-8',hstrip=False)
    >>> print(astr(a, 1, pp=True))
    [['1.4']
     ['2.4']
     ['3.4']
     ['4.4']]

    # openpyxl
    >>> filename = 'test_readexcel.xlsx'
    >>> print(astr(xread(filename, skip=1), 1, pp=True))
    [['1.1' '1.2' '1.3' '1.4']
     ['2.1' '2.2' '2.3' '2.4']
     ['3.1' '3.2' '3.3' '3.4']
     ['4.1' '4.2' '4.3' '4.4']]
    >>> print(astr(xread(filename, skip=1, nc=[2], squeeze=True), 1, pp=True))
    ['1.3' '2.3' '3.3' '4.3']
    >>> print(astr(xread(filename, skip=1, cname=['head1', 'head2']),
    ...            1, pp=True))
    [['1.1' '1.2']
     ['2.1' '2.2']
     ['3.1' '3.2']
     ['4.1' '4.2']]
    >>> a, sa = xread(filename, sheet='Sheet3', nc=[1], snc=[0, 2], skip=1,
    ...               squeeze=True)
    >>> print(astr(a, 1, pp=True))
    ['1.2' '2.2' '3.2' '4.2']
    >>> print(sa)
    [['name1' 'name5']
     ['name2' 'name6']
     ['name3' 'name7']
     ['name4' 'name8']]
    >>> a, sa = xread(filename, sheet=2, cname='head2', snc=[0, 2], skip=1,
    ...               squeeze=True)
    >>> print(astr(a, 1, pp=True))
    ['1.2' '2.2' '3.2' '4.2']
    >>> print(sa)
    [['name1' 'name5']
     ['name2' 'name6']
     ['name3' 'name7']
     ['name4' 'name8']]
    >>> a, sa = xread(filename, sheet='Sheet2', cname=['head2', 'head4'],
    ...               snc=[0, 2], skip=1, fill=True, fill_value=-9,
    ...               sfill_value='-8')
    >>> print(astr(a, 1, pp=True))
    [['-9.0' ' 1.4']
     [' 2.2' ' 2.4']
     [' 3.2' ' 3.4']
     [' 4.2' ' 4.4']]
    >>> print(sa)
    [['1.1' '1.3']
     ['2.1' '2.3']
     ['3.1' '-8']
     ['4.1' '4.3']]
    >>> a, sa = xread(filename, sheet='Sheet2', cname=['head2', 'head4'],
    ...               snc=[0, 2], skip=1, header=True)
    >>> print(a)
    ['head2' 'head4']
    >>> print(sa)
    ['head1' 'head3']
    >>> print(xread(filename, sheet='Sheet2', skip=1, header=True))
    ['head1' 'head2' 'head3' 'head4']
    >>> a, sa = xread(filename, sheet='Sheet2', cname=[' head2', 'head4'],
    ...               snc=[0, 2], skip=1, fill=True, fill_value=-9,
    ...               sfill_value='-8',hstrip=False)
    >>> print(astr(a, 1, pp=True))
    [['1.4']
     ['2.4']
     ['3.4']
     ['4.4']]


    License
    -------
    This file is part of the JAMS Python package, distributed under the MIT
    License.

    Copyright (c) 2017-2019 Matthias Cuntz - mc (at) macu (dot) de

    Permission is hereby granted, free of charge, to any person obtaining a
    copy of this software and associated documentation files (the "Software"),
    to deal in the Software without restriction, including without limitation
    the rights to use, copy, modify, merge, publish, distribute, sublicense,
    and/or sell copies of the Software, and to permit persons to whom the
    Software is furnished to do so, subject to the following conditions:

    The above copyright notice and this permission notice shall be included in
    all copies or substantial portions of the Software.

    THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
    IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
    FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
    AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
    LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING
    FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER
    DEALINGS IN THE SOFTWARE.


    History
    -------
    Written,  Matthias Cuntz, Feb 2017 - modified fsread
    Modified, Matthias Cuntz, Nov 2017 - file->infile, hstrip
              Matthias Cuntz, Feb 2019 - NA -> NaN, i.e. R to Python convention
              Matthias Cuntz, Jul 2020 - use openpyxl for xlsx
                                       - flake8
    """
    warn('The function xread is deprecated from JAMS. Use module pyjams.',
         category=DeprecationWarning)
    # Input error
    if (nc == -1) and (snc == -1):
        estr  = 'nc and snc must be numbers or list of indices;'
        estr += ' -1 means read the rest. nc and snc cannot both be -1.'
        raise IOError(estr)

    # If no float or string column given, then read all as float
    if (nc == 0) and (cname is None) and (snc == 0) and (sname is None):
        nc = -1

    # Open file
    ixlrd = False
    try:
        import xlrd
        wb = xlrd.open_workbook(infile)
        ixlrd = True
    except (ModuleNotFoundError, xlrd.biffh.XLRDError):
        try:
            import openpyxl
            wb = openpyxl.open(infile, read_only=True, data_only=True)
        except ModuleNotFoundError:
            raise IOError('Cannot open file (1) '+infile)
    except IOError:
        raise IOError('Cannot open file (2) '+infile)

    # Get Sheet
    if sheet is None:
        if ixlrd:
            sh = wb.sheet_by_index(0)
        else:
            sh = wb[wb.sheetnames[0]]
    else:
        if type(sheet) is str:
            try:
                if ixlrd:
                    sh = wb.sheet_by_name(sheet)
                else:
                    sh = wb[sheet]
            except IOError:
                raise IOError('Sheet '+sheet+' not in Excel file '+infile)
        else:
            if ixlrd:
                if sheet > wb.nsheets:
                    estr  = 'Error extracting sheet ' + str(sheet)
                    estr += '. Only ' + str(wb.nsheets) + ' in Excel file '
                    estr += infile
                    raise IOError(estr)
                sh = wb.sheet_by_index(sheet)
            else:
                if sheet > len(wb.sheetnames):
                    estr  = 'Error extracting sheet ' + str(sheet)
                    estr += '. Only ' + str(len(wb.sheetnames))
                    estr += ' in Excel file ' + infile
                    raise IOError(estr)
                sh = wb[wb.sheetnames[sheet]]

    if ixlrd:
        ncol = sh.ncols
        nrow = sh.nrows - skip
    else:
        ncol = sh.max_column
        nrow = sh.max_row - skip

    # Read header
    if skip > 0:
        head = ['']*(skip-hskip)
        if ixlrd:
            iskip = 0
            while iskip < (skip-hskip):
                head[iskip] = sh.row_values(hskip+iskip)
                iskip += 1
        else:
            iskip = 0
            for rr in sh.iter_rows(iskip, skip-hskip, values_only=True):
                head[iskip] = rr
                iskip += 1
        nhead = iskip

    # Determine indices
    if nc != 0 and cname is not None:
        raise ValueError('nc and cname are mutually exclusive.')
    if snc != 0 and sname is not None:
        raise ValueError('snc and sname are mutually exclusive.')
    if (cname is not None) or (sname is not None):
        # from first header line
        if (skip-hskip) <= 0:
            raise IOError('No header line left for choosing columns by name.')
        hres = head[0]
        if hstrip:
            hres = [ h.strip() for h in hres ]
    if cname is not None:
        if not isinstance(cname, (list, tuple, np.ndarray)):
            cname = [cname]
        if hstrip:
            cname = [ h.strip() for h in cname ]
        nc = []
        for k in range(len(hres)):
            if hres[k] in cname:
                nc.append(k)
    if sname is not None:
        if not isinstance(sname, (list, tuple, np.ndarray)):
            sname = [sname]
        if hstrip:
            sname = [ h.strip() for h in sname ]
        snc = []
        for k in range(len(hres)):
            if hres[k] in sname:
                snc.append(k)
    if (isinstance(nc, (list, tuple, np.ndarray)) and
        isinstance(snc, (list, tuple, np.ndarray))):
        if np.in1d(nc, snc, assume_unique=True).any():
            raise ValueError('float and string indices overlap.')
        iinc  = nc
        iisnc = snc
    elif isinstance(nc, (list, tuple, np.ndarray)):
        iinc   = nc
        iirest = list(range(cskip, ncol))
        for ii in iinc[::-1]:
            del iirest[ii]
        if snc <= -1:
            iisnc = iirest
        else:
            iisnc = iirest[:snc]
    elif isinstance(snc, (list, tuple, np.ndarray)):
        iisnc  = snc
        iirest = list(range(cskip, ncol))
        for ii in iisnc[::-1]:
            del iirest[ii]
        if nc <= -1:
            iinc = iirest
        else:
            iinc = iirest[:nc]
    else:
        # cannot be nc=-1 and snc=-1
        if nc <= -1:
            iisnc = range(cskip, snc)
            iinc  = range(snc, ncol)
        else:
            if snc <= -1:
                iinc  = range(cskip, nc)
                iisnc = range(nc, ncol)
            else:
                # read snc first then nc
                iisnc = range(cskip, snc)
                iinc  = range(snc, snc+nc)
    nnc  = len(iinc)
    nsnc = len(iisnc)
    if nnc > 0:
        miinc = max(iinc)
    else:
        miinc = -1
    if nsnc > 0:
        miisnc = max(iisnc)
    else:
        miisnc = -1
    aiinc = list()
    aiinc.extend(iinc)
    aiinc.extend(iisnc)
    miianc = max(aiinc)

    # Header
    if header:
        if (skip-hskip) > 0:
            if nnc > 0:
                var = []
            if nsnc > 0:
                svar = []
            k = 0
            while k < (skip-hskip):
                hres = head[k]
                if nnc > 0:
                    var.append([ hres[i] for i in iinc ])
                if nsnc > 0:
                    svar.append([ hres[i] for i in iisnc ])
                k += 1
            if k == 1:
                if nnc > 0:
                    var = var[0]
                if nsnc > 0:
                    svar = svar[0]
            if nnc > 0:
                var = np.array(var, dtype=str)
            if nsnc > 0:
                svar = np.array(svar, dtype=str)
        else:
            if not ixlrd:
                wb.close()
            return None
        if transpose:
            if nnc > 0:
                var = var.T
            if nsnc > 0:
                svar = svar.T
        if nnc > 0:
            var = np.squeeze(var)
            if nsnc > 0:
                svar = np.squeeze(svar)
                if not ixlrd:
                    wb.close()
                return var, svar
            else:
                if not ixlrd:
                    wb.close()
                return var
        else:
            svar = np.squeeze(svar)
            if not ixlrd:
                wb.close()
            return svar

    # Values
    if nnc > 0:
        var = np.full((nrow, nnc), np.nan, dtype=float)
        if ixlrd:
            # read by column
            k = 0
            for ix in iinc:
                ilist = sh.col_values(ix, start_rowx=skip, end_rowx=sh.nrows)
                ilist = [ 'NaN' if iv == 'NA' else iv for iv in ilist ]
                if fill:
                    ilist = [ fill_value if iv == '' else iv for iv in ilist ]
                var[:, k] = np.array(ilist, dtype=float)
                k += 1
        else:
            # read by row
            k = 0
            # index in openpyxl is 1-based
            for ilist in sh.iter_rows(skip+1, values_only=True):
                ilist = [ ilist[ix] for ix in iinc ]
                ilist = [ 'NaN' if iv == 'NA' else iv for iv in ilist ]
                if fill:
                    ilist = [
                        fill_value if ((iv == '') or (iv is None)) else iv
                        for iv in ilist ]
                var[k, :] = np.array(ilist, dtype=float)
                k += 1
    if nsnc > 0:
        svar = []
        if ixlrd:
            # read by column
            for ix in iisnc:
                ilist = sh.col_values(ix, start_rowx=skip, end_rowx=sh.nrows)
                if fill:
                    ilist = [ sfill_value if iv == '' else iv for iv in ilist ]
                svar.append(ilist)
            svar = np.array(svar, dtype=str).T
        else:
            # read by row
            for ilist in sh.iter_rows(skip+1, values_only=True):
                ilist = [ ilist[ix] for ix in iisnc ]
                if fill:
                    ilist = [
                        sfill_value if ((iv == '') or (iv is None)) else iv
                        for iv in ilist ]
                svar.append(ilist)
            svar = np.array(svar, dtype=str)
    if transpose:
        if nnc > 0:
            var = var.T
        if nsnc > 0:
            svar = svar.T
    if squeeze or reform:
        if nnc > 0:
            var = var.squeeze()
        if nsnc > 0:
            svar = svar.squeeze()

    if not ixlrd:
        wb.close()

    if nnc > 0:
        if nsnc > 0:
            return var, svar
        else:
            return var
    else:
        return svar


def xlsread(*args, **kwargs):
    """
        Wrapper for xread
    """
    return xread(*args, **kwargs)


def xlsxread(*args, **kwargs):
    """
        Wrapper for xread
    """
    return xread(*args, **kwargs)


if __name__ == '__main__':
    import doctest
    doctest.testmod(optionflags=doctest.NORMALIZE_WHITESPACE)

    # from autostring import astr
    # filename = 'test_readexcel.xlsx'
    # print('#1 ', astr(xread(filename, skip=1),
    #                   1, pp=True))
    # print('#2 ', astr(xread(filename, skip=1, nc=[2], squeeze=True),
    #                   1, pp=True))
    # print('#3: ', astr(xread(filename, skip=1, cname=['head1', 'head2']),
    #                    1, pp=True))
    # a, sa = xread(filename, sheet='Sheet3', nc=[1], snc=[0, 2],
    #               skip=1, squeeze=True)
    # print('#4: ', astr(a, 1, pp=True))
    # print('#5: ', sa)
    # a, sa = xread(filename, sheet=2, cname='head2', snc=[0, 2],
    #               skip=1, squeeze=True)
    # print('#6: ', astr(a, 1, pp=True))
    # print('#7: ', sa)
    # a, sa = xread(filename, sheet='Sheet2', cname=['head2', 'head4'],
    #               snc=[0, 2], skip=1, fill=True, fill_value=-9,
    #               sfill_value='-8')
    # print('#8: ', astr(a, 1, pp=True))
    # print('#9: ', sa)
    # a, sa = xread(filename, sheet='Sheet2', cname=['head2', 'head4'],
    #               snc=[0, 2], skip=1, header=True)
    # print('#10: ', a)
    # print('#11: ', sa)
    # print('#12: ', xread(filename, sheet='Sheet2', skip=1, header=True))
